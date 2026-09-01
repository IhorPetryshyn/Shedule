import io
import math
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pulp
import streamlit as st

# ==========================================
# 1. ГЛОБАЛЬНІ НАЛАШТУВАННЯ ТА БАЗА ДАНИХ
# ==========================================
st.set_page_config(page_title="Генератор розкладу (Санітарні норми)", layout="wide")

DAYS = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', "П'ятниця"]
SLOTS = [1, 2, 3, 4, 5, 6, 7, 8]

LESSONS_FILE = "saved_lessons.csv"
RULES_FILE = "saved_rules.csv"

HARD_SUBJECTS = ["математика", "алгебра", "геометрія", "фізика", "хімія", "українська", "іноземна", "англійська", "інформатика"]

def load_data():
    if os.path.exists(LESSONS_FILE):
        st.session_state['lessons_db'] = pd.read_csv(LESSONS_FILE)
    else:
        st.session_state['lessons_db'] = pd.DataFrame(columns=["Клас", "Предмет", "Вчитель", "Кількість годин", "Тиждень"])
        
    if os.path.exists(RULES_FILE):
        st.session_state['rules_db'] = pd.read_csv(RULES_FILE)
    else:
        st.session_state['rules_db'] = pd.DataFrame(columns=["Тип заборони", "Об'єкт (Назва)", "День тижня", "Номер уроку"])

def save_data():
    st.session_state['lessons_db'].to_csv(LESSONS_FILE, index=False)
    st.session_state['rules_db'].to_csv(RULES_FILE, index=False)

if 'lessons_db' not in st.session_state or 'rules_db' not in st.session_state:
    load_data()

# ==========================================
# 2. РОБОТА З EXCEL (ІМПОРТ ТА ШАБЛОН)
# ==========================================
def create_excel_template():
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_lessons_sample = pd.DataFrame([
            {"Клас": "10-А", "Предмет": "Математика", "Вчитель": "Іванов І.І.", "Кількість годин": 4, "Тиждень": "Кожен тиждень"},
            {"Клас": "10-А", "Предмет": "Фізика", "Вчитель": "Петров П.П.", "Кількість годин": 2, "Тиждень": "Кожен тиждень"},
            {"Клас": "11-Б", "Предмет": "Хімія", "Вчитель": "Сидорова С.С.", "Кількість годин": 1, "Тиждень": "Чисельник (Тиждень 1)"},
        ])
        df_lessons_sample.to_excel(writer, sheet_name="Lessons", index=False)

        df_teacher_rules_sample = pd.DataFrame([
            {"Вчитель": "Іванов І.І.", "День тижня": "П'ятниця", "Номер уроку": "8"},
            {"Вчитель": "Петров П.П.", "День тижня": "Понеділок", "Номер уроку": "Весь день"}
        ])
        df_teacher_rules_sample.to_excel(writer, sheet_name="Teacher_Rules", index=False)

        df_school_rules_sample = pd.DataFrame([
            {"Тип заборони": "Вся школа", "Об'єкт (Назва)": "Усі", "День тижня": "П'ятниця", "Номер уроку": "8"},
            {"Тип заборони": "Клас", "Об'єкт (Назва)": "10-А", "День тижня": "Понеділок", "Номер уроку": "1"}
        ])
        df_school_rules_sample.to_excel(writer, sheet_name="School_Rules", index=False)

    output.seek(0)
    return output

def process_uploaded_excel(uploaded_file):
    try:
        xls = pd.ExcelFile(uploaded_file)
        
        if "Lessons" in xls.sheet_names:
            df_lessons = pd.read_excel(xls, sheet_name="Lessons")
            if "Тиждень" not in df_lessons.columns:
                df_lessons["Тиждень"] = "Кожен тиждень"
            df_lessons["Тиждень"] = df_lessons["Тиждень"].fillna("Кожен тиждень")
            st.session_state['lessons_db'] = df_lessons[["Клас", "Предмет", "Вчитель", "Кількість годин", "Тиждень"]].dropna(subset=["Клас", "Предмет", "Вчитель"])
        
        new_rules = []

        if "Teacher_Rules" in xls.sheet_names:
            df_t_rules = pd.read_excel(xls, sheet_name="Teacher_Rules").dropna(subset=["Вчитель"])
            for _, row in df_t_rules.iterrows():
                new_rules.append({
                    "Тип заборони": "Вчитель",
                    "Об'єкт (Назва)": str(row["Вчитель"]).strip(),
                    "День тижня": str(row["День тижня"]).strip(),
                    "Номер уроку": str(row["Номер уроку"]).strip()
                })

        if "School_Rules" in xls.sheet_names:
            df_s_rules = pd.read_excel(xls, sheet_name="School_Rules")
            for _, row in df_s_rules.iterrows():
                r_type = row.get("Тип заборони") or row.get("Тип правила", "Вся школа")
                r_obj = row.get("Об'єкт (Назва)") or row.get("Предмет", "Усі")
                
                if pd.notna(row.get("День тижня")) and pd.notna(row.get("Номер уроку")):
                    new_rules.append({
                        "Тип заборони": str(r_type).strip(),
                        "Об'єкт (Назва)": str(r_obj).strip(),
                        "День тижня": str(row["День тижня"]).strip(),
                        "Номер уроку": str(row["Номер уроку"]).strip()
                    })

        if new_rules:
            st.session_state['rules_db'] = pd.DataFrame(new_rules)

        save_data()
        st.success("✅ Дані з Excel успішно імпортовано!")
        st.rerun()

    except Exception as e:
        st.error(f"❌ Помилка читання Excel-файлу: {e}")

# ==========================================
# 3. МАТЕМАТИЧНЕ ЯДРО ОПТИМІЗАЦІЇ (PuLP)
# ==========================================
def solve_schedule(df_lessons, df_rules, opts):
    lessons = []
    for idx, row in df_lessons.iterrows():
        c_id = str(row["Клас"]).strip()
        subj = str(row["Предмет"]).strip()
        t = str(row["Вчитель"]).strip()
        try:
            count = int(row["Кількість годин"])
        except (ValueError, TypeError):
            count = 1
        w_type = str(row.get("Тиждень", "Кожен тиждень"))
        
        for i in range(count):
            lessons.append((c_id, subj, t, f"{subj}_{idx}_{i}", w_type))
            
    classes_list = sorted(df_lessons["Клас"].dropna().astype(str).unique().tolist())
    teachers_list = sorted(df_lessons["Вчитель"].dropna().astype(str).unique().tolist())

    prob = pulp.LpProblem("School_Sanitary_Schedule_Optimization", pulp.LpMinimize)
    schedule_vars = {}
    penalty_terms = []

    for c_id, subj, t, l_id, wt in lessons:
        for d in DAYS:
            for s in SLOTS:
                var_name = f"g_{c_id}_{l_id}_{d}_{s}".replace(" ", "_").replace("'", "_")
                schedule_vars[(c_id, subj, t, l_id, wt, d, s)] = pulp.LpVariable(var_name, cat='Binary')

    # ОБОВ'ЯЗКОВІ БАЗОВІ ОБМЕЖЕННЯ
    for c_id, subj, t, l_id, wt in lessons:
        prob += pulp.lpSum(schedule_vars[(c_id, subj, t, l_id, wt, d, s)] for d in DAYS for s in SLOTS) == 1

    for c_id in classes_list:
        for d in DAYS:
            for s in SLOTS:
                w1 = [schedule_vars[(c, subj, t, l_id, wt, d, s)] for c, subj, t, l_id, wt in lessons if c == c_id and wt in ["Кожен тиждень", "Чисельник (Тиждень 1)"]]
                prob += pulp.lpSum(w1) <= 1
                w2 = [schedule_vars[(c, subj, t, l_id, wt, d, s)] for c, subj, t, l_id, wt in lessons if c == c_id and wt in ["Кожен тиждень", "Знаменник (Тиждень 2)"]]
                prob += pulp.lpSum(w2) <= 1

    for t_name in teachers_list:
        for d in DAYS:
            for s in SLOTS:
                w1_t = [schedule_vars[(c, subj, t, l_id, wt, d, s)] for c, subj, t, l_id, wt in lessons if t == t_name and wt in ["Кожен тиждень", "Чисельник (Тиждень 1)"]]
                prob += pulp.lpSum(w1_t) <= 1
                w2_t = [schedule_vars[(c, subj, t, l_id, wt, d, s)] for c, subj, t, l_id, wt in lessons if t == t_name and wt in ["Кожен тиждень", "Знаменник (Тиждень 2)"]]
                prob += pulp.lpSum(w2_t) <= 1

    # ОБРОБКА ВСІХ ТИПІВ ЗАБОРОН (Вчитель, Клас, Вся школа)
    for _, row in df_rules.dropna(subset=["Тип заборони", "День тижня", "Номер уроку"]).iterrows():
        r_type = str(row["Тип заборони"]).strip()
        r_obj = str(row["Об'єкт (Назва)"]).strip()
        r_day = str(row["День тижня"]).strip()
        r_slot_val = str(row["Номер уроку"]).strip()
        
        slots_to_ban = SLOTS if r_slot_val.lower() == "весь день" else [int(r_slot_val)]
        
        for s in slots_to_ban:
            if r_type == "Вчитель":
                vars_to_ban = [schedule_vars[(c, subj, t, l_id, wt, r_day, s)] for c, subj, t, l_id, wt in lessons if t == r_obj]
            elif r_type == "Клас":
                vars_to_ban = [schedule_vars[(c, subj, t, l_id, wt, r_day, s)] for c, subj, t, l_id, wt in lessons if c == r_obj]
            elif r_type == "Вся школа":
                vars_to_ban = [schedule_vars[(c, subj, t, l_id, wt, r_day, s)] for c, subj, t, l_id, wt in lessons]
            else:
                vars_to_ban = []
                
            if vars_to_ban:
                prob += pulp.lpSum(vars_to_ban) == 0

    # Без вікон для класів
    for c_id in classes_list:
        for d in DAYS:
            for i in range(1, len(SLOTS)):
                s_curr, s_prev = SLOTS[i], SLOTS[i-1]
                w1_curr = pulp.lpSum([schedule_vars[(c, subj, t, l_id, wt, d, s_curr)] for c, subj, t, l_id, wt in lessons if c == c_id and wt in ["Кожен тиждень", "Чисельник (Тиждень 1)"]])
                w1_prev = pulp.lpSum([schedule_vars[(c, subj, t, l_id, wt, d, s_prev)] for c, subj, t, l_id, wt in lessons if c == c_id and wt in ["Кожен тиждень", "Чисельник (Тиждень 1)"]])
                prob += w1_curr <= w1_prev

    # ==========================================
    # НАЛАШТОВУВАНІ САНІТАРНІ ОБМЕЖЕННЯ (ОПЦІЇ)
    # ==========================================
    
    # Опція 1: Динаміка працездатності (Вівторок/Середа 2-3 уроки для складних предметів)
    if opts.get("sanitary_peak_workload"):
        for c, subj, t, l_id, wt in lessons:
            if any(kw in subj.lower() for kw in HARD_SUBJECTS):
                for d in DAYS:
                    for s in SLOTS:
                        penalty = 0
                        if d in ["Понеділок", "П'ятниця"]: penalty += 3
                        if s in [1, 7, 8]: penalty += 4
                        elif s in [2, 3, 4] and d in ["Вівторок", "Середа"]: penalty -= 2
                        
                        if penalty > 0:
                            p_var = pulp.LpVariable(f"p_hard_{l_id}_{d}_{s}", cat='Continuous')
                            prob += p_var >= penalty * schedule_vars[(c, subj, t, l_id, wt, d, s)]
                            penalty_terms.append(p_var)

    # Опція 2: Чергування предметів (Заборона 2 поспіль складних дисциплін)
    if opts.get("alternate_disciplines"):
        for c_id in classes_list:
            for d in DAYS:
                for s in range(1, len(SLOTS)):
                    s1, s2 = SLOTS[s-1], SLOTS[s]
                    h1 = [schedule_vars[(c, subj, t, l_id, wt, d, s1)] for c, subj, t, l_id, wt in lessons if c == c_id and any(kw in subj.lower() for kw in HARD_SUBJECTS)]
                    h2 = [schedule_vars[(c, subj, t, l_id, wt, d, s2)] for c, subj, t, l_id, wt in lessons if c == c_id and any(kw in subj.lower() for kw in HARD_SUBJECTS)]
                    if h1 and h2:
                        prob += pulp.lpSum(h1) + pulp.lpSum(h2) <= 1

    # Опція 3: Контроль здвоєних уроків (Тільки 10-11 кл або лаб./практ./технології)
    if opts.get("enforce_double_lesson_rules"):
        for c_id in classes_list:
            is_senior = any(grade in c_id for grade in ["10", "11", "12"])
            for d in DAYS:
                for s in range(1, len(SLOTS)):
                    s1, s2 = SLOTS[s-1], SLOTS[s]
                    class_subjs = set(subj for c, subj, t, l_id, wt in lessons if c == c_id)
                    for subj_name in class_subjs:
                        is_allowed_double = is_senior or any(kw in subj_name.lower() for kw in ["лабораторна", "практична", "трудове", "технології"])
                        if not is_allowed_double:
                            v1 = [schedule_vars[(c, subj, t, l_id, wt, d, s1)] for c, subj, t, l_id, wt in lessons if c == c_id and subj == subj_name]
                            v2 = [schedule_vars[(c, subj, t, l_id, wt, d, s2)] for c, subj, t, l_id, wt in lessons if c == c_id and subj == subj_name]
                            if v1 and v2: prob += pulp.lpSum(v1) + pulp.lpSum(v2) <= 1

    # Опція 4: Обмеження Фізкультури (Не 1-м та не 7-8 уроками)
    if opts.get("pe_slot_restrictions"):
        for c, subj, t, l_id, wt in lessons:
            if "фізкультура" in subj.lower():
                for d in DAYS:
                    prob += schedule_vars[(c, subj, t, l_id, wt, d, 1)] == 0
                    prob += schedule_vars[(c, subj, t, l_id, wt, d, 7)] == 0
                    prob += schedule_vars[(c, subj, t, l_id, wt, d, 8)] == 0

    # Опція 5: Максимальна кількість уроків на день (Загальний ліміт)
    if opts.get("enable_max_daily_lessons"):
        max_daily = opts.get("max_daily_lessons", 7)
        for c_id in classes_list:
            for d in DAYS:
                prob += pulp.lpSum(schedule_vars[(c, subj, t, l_id, wt, d, s)] for c, subj, t, l_id, wt in lessons if c == c_id for s in SLOTS) <= max_daily

    # Опція 6: Обмеження для початкової школи (1-4 класи)
    if opts.get("enable_screen_time_limits"):
        for c_id in classes_list:
            digits = "".join([ch for ch in c_id if ch.isdigit()])
            if digits:
                g_num = int(digits)
                limit = 7
                if g_num == 1: limit = 4
                elif 2 <= g_num <= 4: limit = 5
                
                if limit < 7:
                    for d in DAYS:
                        prob += pulp.lpSum(schedule_vars[(c, subj, t, l_id, wt, d, s)] for c, subj, t, l_id, wt in lessons if c == c_id for s in SLOTS) <= limit

    if penalty_terms:
        prob += pulp.lpSum(penalty_terms)

    prob.solve(pulp.PULP_CBC_CMD(msg=False))
    return pulp.LpStatus[prob.status], schedule_vars, lessons, classes_list, teachers_list

# ==========================================
# 4. ІНТЕРФЕЙС STREAMLIT
# ==========================================
st.title("🏫 Генератор шкільного розкладу")

tab1, tab2, tab3, tab4 = st.tabs([
    "📋 Навантаження та Excel", 
    "🚫 Заборони (Вчителі, Класи, Школа)", 
    "⚙️ Налаштування та Генерація", 
    "ℹ️ Санітарний регламент (Довідка)"
])

# --- ВКЛАДКА 1: НАВАНТАЖЕННЯ І EXCEL ---
with tab1:
    with st.expander("📁 Імпорт даних з Excel та завантаження шаблону", expanded=False):
        col_ex1, col_ex2 = st.columns([1, 1])
        with col_ex1:
            st.markdown("**1. Завантажте готовий шаблон Excel:**")
            template_bytes = create_excel_template()
            st.download_button(
                label="📥 Скачати шаблон (input_data.xlsx)",
                data=template_bytes,
                file_name="input_data_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        with col_ex2:
            st.markdown("**2. Завантажте заповнений файл Excel:**")
            uploaded_file = st.file_uploader("Виберіть .xlsx файл", type=["xlsx"], label_visibility="collapsed")
            if uploaded_file is not None:
                if st.button("🔄 Зчитати та перезаписати базу", type="primary", use_container_width=True):
                    process_uploaded_excel(uploaded_file)
                    
    st.markdown("---")
    st.subheader("1. Швидке додавання предмета")
    with st.form("quick_add_form", clear_on_submit=True):
        f_col1, f_col2, f_col3, f_col4, f_col5 = st.columns(5)
        c_id = f_col1.text_input("Клас (напр: 10-A):")
        subj = f_col2.text_input("Предмет:")
        teacher = f_col3.text_input("Вчитель:")
        hours = f_col4.number_input("Годин на тиждень:", min_value=1, max_value=10, value=1)
        week_select = f_col5.selectbox("Тиждень:", ["Кожен тиждень", "Чисельник (Тиждень 1)", "Знаменник (Тиждень 2)"])
        
        if st.form_submit_button("➕ Додати у базу"):
            if c_id and subj and teacher:
                new_row = pd.DataFrame([{"Клас": c_id, "Предмет": subj, "Вчитель": teacher, "Кількість годин": hours, "Тиждень": week_select}])
                st.session_state['lessons_db'] = pd.concat([st.session_state['lessons_db'], new_row], ignore_index=True)
                save_data()
                st.rerun()

    st.subheader("2. Інтерактивна таблиця предметів")
    edited_lessons = st.data_editor(st.session_state['lessons_db'], use_container_width=True, num_rows="dynamic")
    if not edited_lessons.equals(st.session_state['lessons_db']):
        st.session_state['lessons_db'] = edited_lessons
        save_data()

# --- ВКЛАДКА 2: ВСІ ТИПИ ЗАБОРОН ---
with tab2:
    st.subheader("🚫 Керування заборонами розкладу")
    
    ban_tab1, ban_tab2, ban_tab3, ban_tab4 = st.tabs([
        "👨‍🏫 Заборони Вчителів (Сітка)", 
        "🏫 Заборони Класів (Сітка)", 
        "🏢 Загальношкільні заборони",
        "📋 Повна таблиця правил"
    ])

    # 1. Заборони Вчителів (Графічна сітка)
    with ban_tab1:
        st.markdown("##### 📅 Вільні/зайняті уроки вчителів")
        active_teachers = sorted(st.session_state['lessons_db']["Вчитель"].dropna().astype(str).unique().tolist())
        if active_teachers:
            selected_teacher = st.selectbox("Виберіть вчителя:", active_teachers, key="sel_t")
            grid_data = {day: [False] * len(SLOTS) for day in DAYS}
            
            t_rules = st.session_state['rules_db'][
                (st.session_state['rules_db']["Тип заборони"] == "Вчитель") & 
                (st.session_state['rules_db']["Об'єкт (Назва)"] == selected_teacher)
            ]
            for _, r in t_rules.iterrows():
                d, s_val = r["День тижня"], str(r["Номер уроку"]).strip()
                if d in grid_data:
                    if s_val.lower() == "весь день": grid_data[d] = [True] * len(SLOTS)
                    elif s_val.isdigit() and 0 <= int(s_val) - 1 < len(SLOTS): grid_data[d][int(s_val) - 1] = True

            df_t_grid = pd.DataFrame(grid_data, index=[f"Урок №{s}" for s in SLOTS])
            st.caption("☑️ Позначено галочкою = Вчитель НЕ МОЖЕ проводити урок у цей час")
            edited_t_grid = st.data_editor(df_t_grid, use_container_width=True, key=f"grid_t_{selected_teacher}")

            if st.button(f"💾 Зберегти графік для {selected_teacher}", type="primary"):
                st.session_state['rules_db'] = st.session_state['rules_db'][
                    ~((st.session_state['rules_db']["Тип заборони"] == "Вчитель") & 
                      (st.session_state['rules_db']["Об'єкт (Назва)"] == selected_teacher))
                ]
                new_r = []
                for d in DAYS:
                    if all(edited_t_grid.loc[f"Урок №{s}", d] for s in SLOTS):
                        new_r.append({"Тип заборони": "Вчитель", "Об'єкт (Назва)": selected_teacher, "День тижня": d, "Номер уроку": "Весь день"})
                    else:
                        for s in SLOTS:
                            if edited_t_grid.loc[f"Урок №{s}", d]:
                                new_r.append({"Тип заборони": "Вчитель", "Об'єкт (Назва)": selected_teacher, "День тижня": d, "Номер уроку": str(s)})
                if new_r:
                    st.session_state['rules_db'] = pd.concat([st.session_state['rules_db'], pd.DataFrame(new_r)], ignore_index=True)
                save_data()
                st.toast("✅ Графік вчителя збережено!")
        else:
            st.info("Спочатку додайте вчителів у вкладці 'Навантаження'.")

    # 2. Заборони Класів (Графічна сітка)
    with ban_tab2:
        st.markdown("##### 📅 Дні та уроки, коли клас НЕ ВЧИТЬСЯ")
        active_classes = sorted(st.session_state['lessons_db']["Клас"].dropna().astype(str).unique().tolist())
        if active_classes:
            selected_class = st.selectbox("Виберіть клас:", active_classes, key="sel_c")
            grid_c_data = {day: [False] * len(SLOTS) for day in DAYS}
            
            c_rules = st.session_state['rules_db'][
                (st.session_state['rules_db']["Тип заборони"] == "Клас") & 
                (st.session_state['rules_db']["Об'єкт (Назва)"] == selected_class)
            ]
            for _, r in c_rules.iterrows():
                d, s_val = r["День тижня"], str(r["Номер уроку"]).strip()
                if d in grid_c_data:
                    if s_val.lower() == "весь день": grid_c_data[d] = [True] * len(SLOTS)
                    elif s_val.isdigit() and 0 <= int(s_val) - 1 < len(SLOTS): grid_c_data[d][int(s_val) - 1] = True

            df_c_grid = pd.DataFrame(grid_c_data, index=[f"Урок №{s}" for s in SLOTS])
            st.caption("☑️ Позначено галочкою = У класу НЕ ПЛАНУЄТЬСЯ урок у цей час")
            edited_c_grid = st.data_editor(df_c_grid, use_container_width=True, key=f"grid_c_{selected_class}")

            if st.button(f"💾 Зберегти графік для {selected_class}", type="primary"):
                st.session_state['rules_db'] = st.session_state['rules_db'][
                    ~((st.session_state['rules_db']["Тип заборони"] == "Клас") & 
                      (st.session_state['rules_db']["Об'єкт (Назва)"] == selected_class))
                ]
                new_r = []
                for d in DAYS:
                    if all(edited_c_grid.loc[f"Урок №{s}", d] for s in SLOTS):
                        new_r.append({"Тип заборони": "Клас", "Об'єкт (Назва)": selected_class, "День тижня": d, "Номер уроку": "Весь день"})
                    else:
                        for s in SLOTS:
                            if edited_c_grid.loc[f"Урок №{s}", d]:
                                new_r.append({"Тип заборони": "Клас", "Об'єкт (Назва)": selected_class, "День тижня": d, "Номер уроку": str(s)})
                if new_r:
                    st.session_state['rules_db'] = pd.concat([st.session_state['rules_db'], pd.DataFrame(new_r)], ignore_index=True)
                save_data()
                st.toast("✅ Графік класу збережено!")
        else:
            st.info("Спочатку додайте класи у вкладці 'Навантаження'.")

    # 3. Загальношкільні заборони
    with ban_tab3:
        st.markdown("##### 🏢 Заборона уроків для ВСІЄЇ школи (наприклад, 8-й урок у П'ятницю / Педрада)")
        with st.form("school_ban_form", clear_on_submit=True):
            sb_col1, sb_col2, sb_col3 = st.columns(3)
            sb_day = sb_col1.selectbox("День тижня:", DAYS)
            sb_slot = sb_col2.selectbox("Номер уроку:", ["Весь день", "1", "2", "3", "4", "5", "6", "7", "8"])
            sb_desc = sb_col3.text_input("Об'єкт / Примітка:", value="Усі")
            
            if st.form_submit_button("➕ Додати загальношкільну заборону"):
                new_s_rule = pd.DataFrame([{"Тип заборони": "Вся школа", "Об'єкт (Назва)": sb_desc, "День тижня": sb_day, "Номер уроку": sb_slot}])
                st.session_state['rules_db'] = pd.concat([st.session_state['rules_db'], new_s_rule], ignore_index=True)
                save_data()
                st.rerun()

    # 4. Повна редагована таблиця
    with ban_tab4:
        st.markdown("##### 📊 Загальний реєстр усіх правил та обмежень")
        edited_rules = st.data_editor(
            st.session_state['rules_db'], 
            use_container_width=True, 
            num_rows="dynamic",
            column_config={
                "Тип заборони": st.column_config.SelectboxColumn(options=["Вчитель", "Клас", "Вся школа"]),
                "День тижня": st.column_config.SelectboxColumn(options=DAYS),
                "Номер уроку": st.column_config.SelectboxColumn(options=["Весь день", "1", "2", "3", "4", "5", "6", "7", "8"])
            }
        )
        if not edited_rules.equals(st.session_state['rules_db']):
            st.session_state['rules_db'] = edited_rules
            save_data()

# --- ВКЛАДКА 3: ГЕНЕРАЦІЯ ТА САНІТАРНІ ОПЦІЇ ---
with tab3:
    st.subheader("⚙️ Санітарні норми та генерація розкладу")
    st.markdown("Виберіть **галочками**, які санітарно-гігієнічні правила повинен врахувати алгоритм:")

    col_s1, col_s2 = st.columns(2)
    with col_s1:
        st.markdown("##### 🧠 Динаміка працездатності та дисципліни")
        opt_peak = st.checkbox("1. Оптимізувати складні предмети на Вівторок/Середу (2–3 уроки)", value=True, help="Математика, мови, фізика та хімія ставляться у години найвищої працездатності")
        opt_alt = st.checkbox("2. Заборонити 2 поспіль складних предмети (чергувати з легкими)", value=True, help="Не допускає послідовного проведення двох складних STEM/мовних уроків")
        opt_double = st.checkbox("3. Обмежити здвоєні уроки (тільки 10–11 кл або лаб./технології)", value=True, help="Забороняє спарені уроки у 5–9 класах за винятком трудового/практик")

    with col_s2:
        st.markdown("##### 🏃 Рухова активність та навантаження")
        opt_pe = st.checkbox("4. Не ставити Фізкультуру 1-м або 7–8 уроками", value=True, help="Розподіляє уроки фізкультури у середині навчального дня")
        
        opt_max_daily_enabled = st.checkbox("5. Обмежити максимальну кількість уроків на день", value=True)
        if opt_max_daily_enabled:
            opt_max_daily_val = st.number_input("Максимум уроків на день для старших класів:", min_value=4, max_value=8, value=7)
        else:
            opt_max_daily_val = 8
            
        opt_screen_limits = st.checkbox("6. Враховувати норми для початкової школи (1–4 класи)", value=True, help="Автоматично обмежує денну кількість занять для 1–4 класів")

    opts_dict = {
        "sanitary_peak_workload": opt_peak,
        "alternate_disciplines": opt_alt,
        "enforce_double_lesson_rules": opt_double,
        "pe_slot_restrictions": opt_pe,
        "enable_max_daily_lessons": opt_max_daily_enabled,
        "max_daily_lessons": opt_max_daily_val,
        "enable_screen_time_limits": opt_screen_limits
    }

    st.markdown("---")
    col_gen1, col_gen2 = st.columns([2, 1])
    with col_gen1:
        run_calc = st.button("🚀 ЗАПУСТИТИ АВТОМАТИЧНИЙ РОЗРАХУНОК", type="primary", use_container_width=True)
    with col_gen2:
        if 'generated_schedule' in st.session_state and not st.session_state['generated_schedule'].empty:
            if st.button("🔄 Скинути розклади", use_container_width=True):
                del st.session_state['generated_schedule']
                st.rerun()

    if run_calc:
        df_lessons_clean = st.session_state['lessons_db'].dropna(subset=["Клас", "Предмет", "Вчитель"])
        with st.spinner("Оптимізація розкладу за обраними правилами..."):
            status, schedule_vars, lessons_data, classes_list, teachers_list = solve_schedule(df_lessons_clean, st.session_state['rules_db'], opts_dict)
            
            if status == "Optimal":
                st.balloons()
                st.success("🎉 Сформовано розклад відповідно до обраних налаштувань!")
                
                weekly_rows = []
                for day in DAYS:
                    for slot in SLOTS:
                        for class_id in classes_list:
                            w1_item, w2_item = "—", "—"
                            for c, subj, t, l_id, wt in lessons_data:
                                if str(c) == str(class_id) and schedule_vars[(c, subj, t, l_id, wt, day, slot)].varValue is not None:
                                    if schedule_vars[(c, subj, t, l_id, wt, day, slot)].varValue > 0.9:
                                        if wt == "Кожен тиждень":
                                            w1_item = f"{subj} ({t})"
                                            w2_item = f"{subj} ({t})"
                                        elif wt == "Чисельник (Тиждень 1)":
                                            w1_item = f"🔼 {subj} ({t})"
                                        elif wt == "Знаменник (Тиждень 2)":
                                            w2_item = f"🔽 {subj} ({t})"
                            display_text = w1_item if w1_item == w2_item else f"{w1_item} / {w2_item}"
                            weekly_rows.append({"Клас": class_id, "День тижня": day, "Номер уроку": slot, "Розклад (Чисельник / Знаменник)": display_text})
                
                st.session_state['generated_schedule'] = pd.DataFrame(weekly_rows)
            else:
                st.error("❌ Не вдалося побудувати розклад. Спробуйте вимкнути деякі санітарні опції або зменшити кількість заборон у вкладці 'Заборони'.")

    # Відображення, Swap та Експорт
    if 'generated_schedule' in st.session_state and not st.session_state['generated_schedule'].empty:
        df_sched = st.session_state['generated_schedule']
        
        st.markdown("---")
        st.subheader("✏️ Інструмент швидкого переміщення (Swap)")
        
        with st.expander("🔄 Обміняти два уроки місцями для класу", expanded=False):
            swap_col1, swap_col2, swap_col3 = st.columns(3)
            sel_class = swap_col1.selectbox("Виберіть клас:", sorted(df_sched["Клас"].unique()))
            
            with swap_col2:
                day1 = st.selectbox("День 1:", DAYS, key="d1")
                slot1 = st.selectbox("Урок 1:", SLOTS, key="s1")
            
            with swap_col3:
                day2 = st.selectbox("День 2:", DAYS, key="d2")
                slot2 = st.selectbox("Урок 2:", SLOTS, key="s2")
            
            if st.button("🔄 Обміняти уроки", use_container_width=True):
                idx1 = df_sched[(df_sched["Клас"] == sel_class) & (df_sched["День тижня"] == day1) & (df_sched["Номер уроку"] == slot1)].index
                idx2 = df_sched[(df_sched["Клас"] == sel_class) & (df_sched["День тижня"] == day2) & (df_sched["Номер уроку"] == slot2)].index
                
                if not idx1.empty and not idx2.empty:
                    val1 = df_sched.loc[idx1[0], "Розклад (Чисельник / Знаменник)"]
                    val2 = df_sched.loc[idx2[0], "Розклад (Чисельник / Знаменник)"]
                    
                    df_sched.loc[idx1[0], "Розклад (Чисельник / Знаменник)"] = val2
                    df_sched.loc[idx2[0], "Розклад (Чисельник / Знаменник)"] = val1
                    
                    st.session_state['generated_schedule'] = df_sched
                    st.toast("✅ Уроки успішно поміняно місцями!")
                    st.rerun()

        # ДЕТЕКТОР КОНФЛІКТІВ ВЧИТЕЛІВ
        teacher_conflicts = []
        temp_records = []
        for _, row in df_sched.iterrows():
            val = row["Розклад (Чисельник / Знаменник)"]
            if val != "—":
                parts = val.split(" / ")
                for p in parts:
                    if "(" in p and ")" in p:
                        t_name = p.split("(")[1].split(")")[0].strip()
                        temp_records.append({"Вчитель": t_name, "День": row["День тижня"], "Слот": row["Номер уроку"], "Клас": row["Клас"]})
        
        if temp_records:
            check_df = pd.DataFrame(temp_records)
            duplicates = check_df[check_df.duplicated(subset=["Вчитель", "День", "Слот"], keep=False)]
            if not duplicates.empty:
                for (t_name, d, s), group in duplicates.groupby(["Вчитель", "День", "Слот"]):
                    classes_str = ", ".join(group["Клас"].unique())
                    teacher_conflicts.append(f"⚠️ **{t_name}** має накладку в **{d}**, **{s}-й урок** у класах: {classes_str}")

        if teacher_conflicts:
            st.warning("Виявлено конфлікти після ручних правок:")
            for conf in teacher_conflicts: st.write(conf)

        st.subheader("📅 Розклад за днями тижня (Вся школа):")
        day_tabs = st.tabs([f"🗓️ {day}" for day in DAYS])
        
        for idx, day_name in enumerate(DAYS):
            with day_tabs[idx]:
                df_day = df_sched[df_sched["День тижня"] == day_name]
                pivot_day = df_day.pivot(index="Номер уроку", columns="Клас", values="Розклад (Чисельник / Знаменник)")
                
                edited_pivot = st.data_editor(pivot_day, use_container_width=True, key=f"editor_day_{day_name}")
                
                for class_col in edited_pivot.columns:
                    for slot_row in edited_pivot.index:
                        val = edited_pivot.loc[slot_row, class_col]
                        mask = (df_sched["День тижня"] == day_name) & (df_sched["Номер уроку"] == slot_row) & (df_sched["Клас"] == class_col)
                        df_sched.loc[mask, "Розклад (Чисельник / Знаменник)"] = val
        
        st.session_state['generated_schedule'] = df_sched

        # ЕКСПОРТ В EXCEL
        wb = Workbook()
        wb.remove(wb.active)
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Arial", size=10)
        fill_header = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
        thin_border = Border(left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'), top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'))

        for class_id in sorted(df_sched["Клас"].unique()):
            ws = wb.create_sheet(title=f"Клас {class_id}")
            ws.sheet_view.showGridLines = True
            headers = ["День тижня", "Номер уроку", "Розклад (Чисельник / Знаменник)"]
            ws.append(headers)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = font_header; cell.fill = fill_header; cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            class_data = df_sched[df_sched["Клас"] == class_id]
            for _, row_data in class_data.iterrows():
                ws.append([row_data["День тижня"], f"Урок №{row_data['Номер уроку']}", row_data["Розклад (Чисельник / Знаменник)"]])
                curr_row = ws.max_row
                for col_num in range(1, 4):
                    c = ws.cell(row=curr_row, column=col_num)
                    c.font = font_body; c.border = thin_border
                    c.alignment = Alignment(horizontal="center" if col_num <= 2 else "left", vertical="center")

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        st.markdown("---")
        st.download_button(
            label="📥 СКАЧАТИ ВІДКОРИГОВАНИЙ РОЗКЛАД В EXCEL (.xlsx)", 
            data=excel_buffer, 
            file_name="school_schedule_edited.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            use_container_width=True
        )

# --- ВКЛАДКА 4: ДОВІДНИК САНІТАРНОГО РЕГЛАМЕНТУ ---
with tab4:
    st.subheader("📜 Вимоги Санітарного регламенту для ЗЗСО")
    col_r1, col_r2 = st.columns(2)
    with col_r1:
        st.markdown("""
        **1. Тривалість занять:**
        * **1 клас:** 35 хвилин
        * **2–4 класи:** 40 хвилин
        * **5–12 класи:** 45 хвилин
        
        **2. Тривалість перерв:**
        * Мала перерва: не менше **10 хвилин**.
        * Велика перерва (харчування/відпочинок): **20–30 хвилин** (або дві по 20 хвилин).
        """)
    with col_r2:
        st.markdown("""
        **3. Робота з екранами (ТЗН) під час онлайн-уроку:**
        * **1 класи:** не більше 10 хвилин
        * **2–4 класи:** не більше 15 хвилин
        * **5–7 класи:** не більше 20 хвилин
        * **8–9 класи:** 20–25 хвилин
        * **10–12 класи:** 25–30 хвилин (на 1-й години), 15–20 хвилин (на 2-й години).
        """)
